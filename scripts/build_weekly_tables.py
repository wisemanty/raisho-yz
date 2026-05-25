#!/usr/bin/env python3
"""Build RAISHO weekly Youzan operating tables from a core order-item export."""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INVALID_STATUS_PATTERNS = ["关闭", "取消", "未支付", "待付款", "退款成功", "全额退款"]

PRODUCT_LIFECYCLE_NOTES = {
    "玉乃光黑标": "黑标为2026-05-10新上预售，约2026-05-31到货；预售期按高信任信号处理，不用复购缺失扣分。",
    "今治毛巾": "今治毛巾为2026-05-10新上预售，约2026-05-31到货；预售期重点看转介绍和收货反馈。",
    "今治浴巾": "今治浴巾为2026-05-10新上预售，约2026-05-31到货；预售期重点看礼赠和家庭使用反馈。",
}


def read_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    with path.open("rb") as fh:
        signature = fh.read(4)
    if suffix in {".xlsx", ".xls"} or signature.startswith(b"PK"):
        return pd.read_excel(path)
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path)


def find_col(columns: Iterable[str], candidates: list[str], required: bool = False) -> str | None:
    cols = [str(c).strip() for c in columns]
    for candidate in candidates:
        for col in cols:
            if col == candidate:
                return col
    for candidate in candidates:
        for col in cols:
            if candidate.lower() in col.lower():
                return col
    if required:
        raise ValueError(f"Missing required column, expected one of: {', '.join(candidates)}")
    return None


def money(series: pd.Series) -> pd.Series:
    cleaned = series.astype(str).str.replace(",", "", regex=False).str.replace("￥", "", regex=False)
    cleaned = cleaned.str.extract(r"(-?\d+(?:\.\d+)?)", expand=False)
    return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def join_unique(values: Iterable, limit: int = 8) -> str:
    seen = []
    for value in values:
        text = clean_text(value)
        if text and text not in seen:
            seen.append(text)
    if len(seen) > limit:
        return " / ".join(seen[:limit]) + f" / +{len(seen) - limit}"
    return " / ".join(seen)


def categorize_product(name: str) -> str:
    text = clean_text(name)
    if "黑标" in text or "至臻" in text:
        return "玉乃光黑标"
    if "玉乃光" in text or "本真" in text or "品鉴" in text:
        return "玉乃光白标"
    if "浴巾" in text:
        return "今治浴巾"
    if "毛巾" in text or "今治" in text:
        return "今治毛巾"
    upper_text = text.upper()
    if "ADDELM" in upper_text or "ADELM" in upper_text or upper_text.startswith("ADD."):
        return "AddElm"
    return "其他"


def effective_mask(df: pd.DataFrame, status_col: str | None) -> pd.Series:
    if not status_col:
        return pd.Series(True, index=df.index)
    status = df[status_col].astype(str)
    mask = pd.Series(True, index=df.index)
    for pattern in INVALID_STATUS_PATTERNS:
        mask &= ~status.str.contains(pattern, na=False)
    return mask


def latest_nonempty(group: pd.DataFrame, col: str | None) -> str:
    if not col:
        return ""
    values = [clean_text(v) for v in group[col].tolist()]
    values = [v for v in values if v]
    return values[-1] if values else ""


def percentile(series: pd.Series, q: float) -> float:
    values = pd.to_numeric(series, errors="coerce").dropna()
    values = values[values > 0]
    if values.empty:
        return 0.0
    return float(values.quantile(q))


def add_user_percentiles(users: pd.DataFrame) -> pd.DataFrame:
    users = users.copy()
    for col, pct_col in [("客单价", "_客单价百分位"), ("累计实付", "_累计实付百分位"), ("有效订单数", "_订单数百分位")]:
        values = pd.to_numeric(users[col], errors="coerce").fillna(0)
        if len(values) <= 1 or values.nunique() <= 1:
            users[pct_col] = 0.5
        else:
            users[pct_col] = values.rank(pct=True, method="average")
    return users


def compute_user_rule_profile(users: pd.DataFrame) -> dict[str, object]:
    return {
        "用户数": int(len(users)),
        "高客单线": round(percentile(users["客单价"], 0.75), 2) if not users.empty else 0,
        "高累计支付线": round(percentile(users["累计实付"], 0.80), 2) if not users.empty else 0,
        "重点累计支付线": round(percentile(users["累计实付"], 0.90), 2) if not users.empty else 0,
        "生成逻辑": "agent按当周用户数据自动校准：客单价前25%、累计实付前20%、重点累计实付前10%，再叠加黑标/预售/复购/分销信号。",
    }


def pct_text(value: float) -> str:
    if pd.isna(value):
        return "未知"
    return f"前{max(1, round((1 - float(value)) * 100))}%"


def classify_user(row: pd.Series, profile: dict[str, object]) -> tuple[str, str, str, str, str, str, str, str, str, str, str]:
    tags = []
    reasons = []
    rules = []
    categories = clean_text(row["购买品类"])
    has_black = "玉乃光黑标" in categories
    has_sake = "玉乃光" in categories
    has_imabari = "今治" in categories
    has_distributor = bool(clean_text(row["分销员归属"]))
    is_repeat = row["有效订单数"] >= 2
    aov_pct = float(row.get("_客单价百分位", 0.5) or 0.5)
    paid_pct = float(row.get("_累计实付百分位", 0.5) or 0.5)
    order_pct = float(row.get("_订单数百分位", 0.5) or 0.5)
    high_money_signal = aov_pct >= 0.75 or paid_pct >= 0.80

    if row["有效订单数"] >= 2:
        tags.append("复购用户")
        reasons.append(f"有效订单数{row['有效订单数']}，已出现复购")
        rules.append("repeat_order")
    else:
        tags.append("尝鲜用户")
    if high_money_signal:
        tags.append("高客单用户")
        reasons.append(f"客单价处于{pct_text(aov_pct)}，累计实付处于{pct_text(paid_pct)}")
        rules.append("dynamic_high_value")
    if has_black or (has_sake and (aov_pct >= 0.60 or paid_pct >= 0.60)):
        tags.append("黑标潜力用户")
        reasons.append("已买黑标或酒类消费信号较强，适合观察白标到黑标的信任迁移")
        rules.append("black_label_potential")
    if has_distributor or is_repeat:
        tags.append("分销潜力用户")
        rules.append("distribution_or_repeat_signal")
    for category, note in PRODUCT_LIFECYCLE_NOTES.items():
        if category in categories:
            reasons.append(note)

    score = 0.0
    if has_black:
        score += 3.0
    if paid_pct >= 0.90:
        score += 2.0
    elif paid_pct >= 0.80:
        score += 1.2
    if aov_pct >= 0.85:
        score += 1.2
    elif aov_pct >= 0.75:
        score += 0.8
    if is_repeat:
        score += 1.2
    if has_distributor:
        score += 0.5
    if has_imabari:
        score += 0.6

    review = "否"
    if has_black or score >= 3.0:
        priority = "P0"
        if has_black:
            action = "预售期一对一确认期待，5/31后收反馈"
            product = "玉乃光黑标；后续承接白标复购/同圈层推荐"
            script = "S02A"
        elif has_sake:
            action = "一对一做白标到黑标的差异解释"
            product = "玉乃光黑标或白标复购组合"
            script = "S02B"
        else:
            action = "高价值用户一对一回访，先问体验再判断升单"
            product = "根据已购品类给升级或复购组合"
            script = "S02C"
        group = "核心品鉴群"
        one_to_one = "必须一对一"
        if not has_black and not is_repeat:
            review = "是：高金额但复购/黑标信号不足，跟进前先看聊天关系"
    elif is_repeat:
        priority = "P1"
        action = "进复购池，发会员/新品权益"
        product = "复购商品或新品组合"
        group = "老客群"
        one_to_one = "建议一对一"
        script = "S03A"
    elif has_imabari or high_money_signal:
        priority = "P1"
        if has_imabari:
            action = "到货后问手感体验，转礼赠/转介绍"
            product = "今治毛巾/浴巾组合"
            script = "S01A"
        else:
            action = "观察升级意向，先问体验再推高阶品"
            product = "白标/黑标或对应高阶组合"
            script = "S02B"
        group = "轻邀请进群或老客群"
        one_to_one = "建议一对一"
    elif has_distributor:
        priority = "P2"
        action = "让分销员做轻触达"
        product = "白标或今治入门款"
        group = "按分销员圈层"
        one_to_one = "可由分销员一对一"
        script = "S04A"
    else:
        priority = "P3"
        action = "内容轻触达，观察复购信号"
        product = "低门槛入门内容"
        group = "暂不强拉群"
        one_to_one = "不必一对一"
        script = "S05A"

    if not reasons:
        reasons.append("只有基础成交信号，暂不做强判断")
    confidence = "高" if len(rules) >= 3 or has_black else "中" if len(rules) >= 2 or is_repeat or high_money_signal else "低"
    rule_text = " / ".join(dict.fromkeys(rules)) or "base_purchase"
    reason_text = "；".join(dict.fromkeys(reasons))

    return " / ".join(tags), priority, action, product, group, one_to_one, script, rule_text, reason_text, confidence, review


def audit_note_summary(audit_note: Path | None) -> pd.DataFrame:
    if not audit_note:
        return pd.DataFrame([{"类型": "后台巡检", "内容": "未提供后台巡检记录"}])
    if not audit_note.exists():
        return pd.DataFrame([{"类型": "后台巡检", "内容": f"未找到巡检记录：{audit_note}"}])

    rows = []
    current_section = ""
    for raw_line in audit_note.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("#"):
            current_section = line.lstrip("#").strip()
            continue
        if line.startswith("| ---"):
            continue
        if line.startswith("|"):
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if cells and cells[0] not in {"模块", "文件", "数据"}:
                rows.append({"类型": current_section or "表格记录", "内容": " | ".join(cells)})
            continue
        rows.append({"类型": current_section or "巡检记录", "内容": line})
    return pd.DataFrame(rows or [{"类型": "后台巡检", "内容": "巡检记录为空"}])


def build_tables(detail_path: Path, output_dir: Path, week_label: str, audit_note: Path | None = None) -> Path:
    raw = read_table(detail_path)
    raw.columns = [str(c).strip() for c in raw.columns]

    uid_col = find_col(raw.columns, ["yz_open_id", "yzOpenId", "有赞客户ID"], required=True)
    nickname_col = find_col(raw.columns, ["客户昵称", "买家昵称", "昵称"])
    phone_col = find_col(raw.columns, ["手机号", "买家手机号", "客户手机号"])
    order_col = find_col(raw.columns, ["订单号", "订单编号"], required=True)
    status_col = find_col(raw.columns, ["订单状态", "交易状态", "状态"])
    pay_time_col = find_col(raw.columns, ["支付时间", "付款时间"])
    order_time_col = find_col(raw.columns, ["下单时间", "订单创建时间", "创建时间"])
    product_col = find_col(raw.columns, ["商品名称", "商品", "商品标题"], required=True)
    distributor_col = find_col(raw.columns, ["分销员", "分销员昵称", "推广员"])
    channel_col = find_col(raw.columns, ["来源渠道", "销售渠道", "订单来源"])
    tag_col = find_col(raw.columns, ["客户标签", "标签"])
    amount_col = find_col(raw.columns, ["商品实收金额", "订单实收", "实收金额", "商品销售金额", "订单金额"], required=True)
    order_amount_col = find_col(raw.columns, ["订单实收", "实收金额", "订单金额"]) or amount_col
    qty_col = find_col(raw.columns, ["商品销售件数", "数量", "件数"])

    df = raw.copy()
    df["_uid"] = df[uid_col].map(clean_text)
    df = df[df["_uid"] != ""].copy()
    df["_order"] = df[order_col].map(clean_text)
    df["_product"] = df[product_col].map(clean_text)
    df["_category"] = df["_product"].map(categorize_product)
    df["_amount"] = money(df[amount_col])
    df["_order_amount"] = money(df[order_amount_col])
    df["_qty"] = money(df[qty_col]) if qty_col else 1
    time_source = pay_time_col or order_time_col
    df["_time"] = pd.to_datetime(df[time_source], errors="coerce") if time_source else pd.NaT
    df["_effective"] = effective_mask(df, status_col)
    eff = df[df["_effective"]].copy()

    order_rows = (
        eff.sort_values("_time")
        .groupby(["_uid", "_order"], dropna=False)
        .agg(
            订单时间=("_time", "min"),
            订单实付=("_order_amount", "max"),
            品类=("_category", lambda x: join_unique(x, 20)),
            商品=("_product", lambda x: join_unique(x, 20)),
        )
        .reset_index()
    )
    order_rows["购买日期"] = pd.to_datetime(order_rows["订单时间"], errors="coerce").dt.date
    session_counts = (
        order_rows.dropna(subset=["购买日期"])
        .groupby("_uid")["购买日期"]
        .nunique()
        .to_dict()
    )

    user_records = []
    for uid, group in eff.sort_values("_time").groupby("_uid"):
        user_orders = order_rows[order_rows["_uid"] == uid].sort_values("订单时间")
        categories = user_orders["品类"].tolist()
        paid = float(user_orders["订单实付"].sum())
        order_count = int(user_orders["_order"].nunique())
        purchase_sessions = int(session_counts.get(uid, order_count))
        aov = paid / order_count if order_count else 0
        rec = {
            "yz_open_id": uid,
            "客户显示": latest_nonempty(group, nickname_col) or uid,
            "历史昵称": join_unique(group[nickname_col]) if nickname_col else "",
            "脱敏手机号": latest_nonempty(group, phone_col),
            "有效订单数": order_count,
            "购买会话数": purchase_sessions,
            "累计实付": round(paid, 2),
            "客单价": round(aov, 2),
            "首单时间": user_orders["订单时间"].min(),
            "最近下单时间": user_orders["订单时间"].max(),
            "首单品类": categories[0] if categories else "",
            "最近品类": categories[-1] if categories else "",
            "购买品类": join_unique(group["_category"], 20),
            "分销员归属": join_unique(group[distributor_col]) if distributor_col else "",
            "来源渠道": join_unique(group[channel_col]) if channel_col else "",
            "客户标签": join_unique(group[tag_col]) if tag_col else "",
        }
        user_records.append(rec)

    users = pd.DataFrame(user_records)
    users = add_user_percentiles(users)
    user_rule_profile = compute_user_rule_profile(users)
    classified = users.apply(lambda r: classify_user(r, user_rule_profile), axis=1, result_type="expand")
    classified.columns = [
        "用户标签", "优先级", "下一步动作", "本周推什么", "是否进群", "是否一对一",
        "话术编号", "命中规则", "判断原因", "建议置信度", "是否需人工复核"
    ]
    users = pd.concat([users, classified], axis=1)
    users["客单价百分位"] = (users["_客单价百分位"] * 100).round(1).astype(str) + "%"
    users["累计实付百分位"] = (users["_累计实付百分位"] * 100).round(1).astype(str) + "%"
    users = users.drop(columns=["_客单价百分位", "_累计实付百分位", "_订单数百分位"])
    users = users.sort_values(
        ["优先级", "累计实付", "有效订单数"], ascending=[True, False, False]
    )

    transitions = []
    path_details = []
    for uid, user_orders in order_rows.sort_values("订单时间").groupby("_uid"):
        seq = []
        for cats in user_orders["品类"].tolist():
            for cat in cats.split(" / "):
                if cat and (not seq or seq[-1] != cat):
                    seq.append(cat)
        user_name = users.loc[users["yz_open_id"] == uid, "客户显示"].iloc[0] if not users.empty else uid
        path_details.append({
            "yz_open_id": uid,
            "客户显示": user_name,
            "品类路径": " -> ".join(seq),
            "订单数": user_orders["_order"].nunique(),
            "累计实付": round(float(user_orders["订单实付"].sum()), 2),
        })
        if len(seq) >= 2:
            transitions.append({
                "起点品类": seq[0],
                "后续品类": seq[1],
                "yz_open_id": uid,
                "客户显示": user_name,
                "金额": float(user_orders["订单实付"].sum()),
            })

    if transitions:
        trans_df = pd.DataFrame(transitions)
        path = (
            trans_df.groupby(["起点品类", "后续品类"])
            .agg(
                迁移人数=("yz_open_id", "nunique"),
                迁移金额=("金额", "sum"),
                典型用户=("客户显示", lambda x: join_unique(x, 5)),
            )
            .reset_index()
        )
        path["信任解释"] = path.apply(lambda r: explain_path(r["起点品类"], r["后续品类"]), axis=1)
        path["下一步商品策略"] = path.apply(lambda r: product_strategy(r["起点品类"], r["后续品类"]), axis=1)
        path["迁移金额"] = path["迁移金额"].round(2)
    else:
        path = pd.DataFrame(columns=["起点品类", "后续品类", "迁移人数", "迁移金额", "典型用户", "信任解释", "下一步商品策略"])

    path_detail_df = pd.DataFrame(path_details)

    if distributor_col:
        dist_records = []
        for dist, group in eff.groupby(distributor_col):
            dist_name = clean_text(dist)
            if not dist_name:
                continue
            related_users = users[users["yz_open_id"].isin(group["_uid"].unique())]
            order_subset = order_rows[order_rows["_uid"].isin(group["_uid"].unique())]
            customer_count = int(group["_uid"].nunique())
            order_count = int(order_subset["_order"].nunique())
            paid = float(order_subset["订单实付"].sum())
            repeat_count = int((related_users["购买会话数"] >= 2).sum()) if "购买会话数" in related_users.columns else int((related_users["有效订单数"] >= 2).sum())
            high_count = int(related_users["用户标签"].astype(str).str.contains("高客单用户", na=False).sum())
            black_count = int(related_users["购买品类"].astype(str).str.contains("玉乃光黑标", na=False).sum())
            dist_records.append({
                "分销员": dist_name,
                "带来客户数": customer_count,
                "有效订单数": order_count,
                "累计实付": round(paid, 2),
                "客单价": round(paid / order_count, 2) if order_count else 0,
                "复购客户数": repeat_count,
                "复购率": f"{repeat_count / customer_count:.1%}" if customer_count else "0.0%",
                "复购口径": "按购买会话数统计；同一用户同一自然日多单合并为1次购买会话。",
                "高客单客户数": high_count,
                "黑标客户数": black_count,
                "主要成交品类": join_unique(group["_category"], 10),
                "分销员口径风险": "按分销员昵称聚合；如分销员改名或重名，需要接入分销员ID/手机号后再合并。",
            })
        distributors = pd.DataFrame(dist_records)
        if not distributors.empty:
            distributors = add_distributor_percentiles(distributors)
            distributors["质量判断"] = distributors.apply(lambda r: distributor_judgement(
                int(r["带来客户数"]),
                float(r["累计实付"]),
                int(r["复购客户数"]),
                int(r["黑标客户数"]),
                int(r["高客单客户数"]),
                float(r["_成交额百分位"]),
                float(r["_客户数百分位"]),
            ), axis=1)
            distributors["判断原因"] = distributors.apply(lambda r: distributor_reason(r), axis=1)
            distributors["培养动作"] = distributors.apply(lambda r: distributor_action(
                int(r["带来客户数"]),
                float(r["累计实付"]),
                int(r["复购客户数"]),
                int(r["黑标客户数"]),
                int(r["高客单客户数"]),
                float(r["_成交额百分位"]),
                float(r["_客户数百分位"]),
            ), axis=1)
            distributors["成交额百分位"] = (distributors["_成交额百分位"] * 100).round(1).astype(str) + "%"
            distributors["客户数百分位"] = (distributors["_客户数百分位"] * 100).round(1).astype(str) + "%"
            distributors = distributors.drop(columns=["_成交额百分位", "_客户数百分位"])
        distributors = distributors.sort_values(["累计实付", "带来客户数"], ascending=[False, False])
    else:
        distributors = pd.DataFrame(columns=["分销员", "带来客户数", "有效订单数", "累计实付", "客单价", "复购客户数", "复购率", "高客单客户数", "黑标客户数", "主要成交品类", "质量判断", "培养动作"])

    actions = users[[
        "优先级", "yz_open_id", "客户显示", "用户标签", "购买品类", "分销员归属",
        "下一步动作", "本周推什么", "是否进群", "是否一对一", "话术编号",
        "命中规则", "判断原因", "建议置信度", "是否需人工复核"
    ]].copy()
    actions["话术变量"] = actions.apply(lambda r: f"品类={r['购买品类']}; 分销员={r['分销员归属']}", axis=1)

    scripts = pd.DataFrame([
        {"话术编号": "S01A", "适用对象": "今治预售/新买家", "话术用途": "到货后体验反馈和转介绍", "话术": "先确认到货节奏，5/31后重点问手感、家用和礼赠反馈；有自然好评再请对方推荐给熟人。"},
        {"话术编号": "S02A", "适用对象": "黑标预售用户", "话术用途": "高信任一对一维护", "话术": "先感谢预售期信任，说明到货节奏；到货后重点问米种、口感和饮用场景，再判断是否进入核心品鉴名单。"},
        {"话术编号": "S02B", "适用对象": "白标/高客单但未买黑标用户", "话术用途": "从白标到黑标的信任升级", "话术": "先问白标体验，再解释黑标和白标的差异，不直接硬推；看对方是否关心米种、风味、限量或礼赠。"},
        {"话术编号": "S02C", "适用对象": "高价值但原因不完全明确用户", "话术用途": "人工复核后一对一", "话术": "先查看关系来源和历史沟通，再决定由老板、核心运营或分销员跟进，避免只因金额高就强推。"},
        {"话术编号": "S03A", "适用对象": "复购用户", "话术用途": "会员/社群承接", "话术": "强调对方已不是第一次购买，邀请进入老客优先名单，后续新品、限量和组合权益先同步。"},
        {"话术编号": "S04A", "适用对象": "分销相关用户/分销员", "话术用途": "分销转化支持", "话术": "给分销员具体客户反馈和一版可转发素材，让分销员先轻触达，再由运营判断是否接入一对一。"},
        {"话术编号": "S05A", "适用对象": "低信号用户", "话术用途": "轻内容触达", "话术": "只发真实使用/饮用内容，不催单；观察再次浏览、咨询、加购、进群或复购信号。"},
    ])
    rule_summary = pd.DataFrame([
        {"项目": key, "说明": value}
        for key, value in user_rule_profile.items()
    ])

    checks = pd.DataFrame([
        {"检查项": "原始文件", "结果": str(detail_path)},
        {"检查项": "原始行数", "结果": len(raw)},
        {"检查项": "含yz_open_id行数", "结果": len(df)},
        {"检查项": "有效明细行数", "结果": len(eff)},
        {"检查项": "去重用户数", "结果": users["yz_open_id"].nunique() if not users.empty else 0},
        {"检查项": "去重订单数", "结果": order_rows["_order"].nunique() if not order_rows.empty else 0},
        {"检查项": "使用时间字段", "结果": time_source or "缺失"},
        {"检查项": "使用金额字段", "结果": amount_col},
        {"检查项": "使用订单金额字段", "结果": order_amount_col},
        {"检查项": "分销员字段", "结果": distributor_col or "缺失"},
        {"检查项": "分销员聚合口径", "结果": "当前按分销员昵称字段聚合；买家客户按yz_open_id去重。若要避免分销员改名/重名，需要补充分销员ID或手机号导出。"},
        {"检查项": "复购统计口径", "结果": "复购客户数/复购率按购买会话数计算；同一用户同一自然日内多次下单合并为1次购买会话，避免规格限制或预售期补量造成复购虚高。"},
        {"检查项": "状态字段", "结果": status_col or "缺失"},
    ])

    output_dir.mkdir(parents=True, exist_ok=True)
    out = output_dir / f"来处有赞周度四张表_{week_label}.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        pd.DataFrame([
            {"项目": "周度", "说明": week_label},
            {"项目": "主键", "说明": "yz_open_id"},
            {"项目": "分销员口径", "说明": "当前核心明细只有分销员昵称/团队，分销员质量表按昵称聚合；买家客户数按 yz_open_id 去重。"},
            {"项目": "复购口径", "说明": "复购客户数/复购率按购买会话数计算：同一用户同一自然日内多单合并为1次购买会话，避免当天补量或预售期追加被误算为复购。"},
            {"项目": "核心框架", "说明": "数据 -> 用户 -> 行为 -> 信任 -> 动作"},
            {"项目": "规则方式", "说明": "优先级和高客单不再使用固定金额门槛；agent先按当周数据分布自动校准，再叠加商品生命周期、复购、分销信号。"},
            {"项目": "预售事实", "说明": "黑标和今治毛巾为2026-05-10新上预售链接，约2026-05-31到货"},
        ]).to_excel(writer, sheet_name="00口径说明", index=False)
        users.to_excel(writer, sheet_name="01用户分层表", index=False)
        path.to_excel(writer, sheet_name="02商品路径表", index=False)
        path_detail_df.to_excel(writer, sheet_name="02路径明细", index=False)
        distributors.to_excel(writer, sheet_name="03分销员质量表", index=False)
        actions.to_excel(writer, sheet_name="04经营动作表", index=False)
        scripts.to_excel(writer, sheet_name="05话术库", index=False)
        rule_summary.to_excel(writer, sheet_name="06规则校准说明", index=False)
        checks.to_excel(writer, sheet_name="07数据质量检查", index=False)
        audit_note_summary(audit_note).to_excel(writer, sheet_name="08后台巡检摘要", index=False)
        format_workbook(writer.book)
    return out


def explain_path(start: str, next_cat: str) -> str:
    if start.startswith("今治") and "玉乃光" in next_cat:
        return "从低门槛生活品迁移到酒类信任，说明日用品可承担破冰。"
    if "白标" in start and "黑标" in next_cat:
        return "从入门酒款升级到高客单预售，说明酒类信任开始加深。"
    if "玉乃光" in start and next_cat.startswith("今治"):
        return "从酒类信任迁移到生活品，适合做礼赠和复购承接。"
    return "存在跨品类复购，需要结合用户明细判断具体信任原因。"


def product_strategy(start: str, next_cat: str) -> str:
    if "黑标" in next_cat:
        return "保留黑标稀缺性，用白标体验和一对一解释承接升级。"
    if next_cat.startswith("今治"):
        return "用触感、礼赠、家庭使用场景推动转介绍。"
    if "玉乃光" in next_cat:
        return "补充产地、米种、饮用场景内容，推动白标复购。"
    return "先观察路径样本，避免过早定义主推逻辑。"


def add_distributor_percentiles(distributors: pd.DataFrame) -> pd.DataFrame:
    distributors = distributors.copy()
    for col, pct_col in [("累计实付", "_成交额百分位"), ("带来客户数", "_客户数百分位")]:
        values = pd.to_numeric(distributors[col], errors="coerce").fillna(0)
        if len(values) <= 1 or values.nunique() <= 1:
            distributors[pct_col] = 0.5
        else:
            distributors[pct_col] = values.rank(pct=True, method="average")
    return distributors


def distributor_judgement(
    customers: int,
    paid: float,
    repeat: int,
    black: int,
    high: int = 0,
    paid_pct: float | None = None,
    customer_pct: float | None = None,
) -> str:
    if black > 0:
        return "重点培养：已经带来黑标/高信任客户"
    if paid_pct is not None and paid_pct >= 0.80:
        return "重点培养：成交额处于本期分销员前列"
    if repeat > 0 and (customer_pct is None or customer_pct >= 0.60):
        return "可重点培养：已有复购且客户基础不弱"
    if high > 0:
        return "可培养：已有高客单客户，需要验证是否可复制"
    if customers > 0:
        return "轻培养：有成交但圈层或复购尚未成立"
    return "暂观察：注册或弱信号"


def distributor_reason(row: pd.Series) -> str:
    reasons = [
        f"带来客户数{row['带来客户数']}，处于{pct_text(float(row.get('_客户数百分位', 0.5)))}",
        f"累计实付{row['累计实付']}，处于{pct_text(float(row.get('_成交额百分位', 0.5)))}",
    ]
    if int(row.get("复购客户数", 0) or 0) > 0:
        reasons.append("已出现复购客户")
    if int(row.get("高客单客户数", 0) or 0) > 0:
        reasons.append("已出现高客单客户")
    if int(row.get("黑标客户数", 0) or 0) > 0:
        reasons.append("已出现黑标客户")
    return "；".join(reasons)


def distributor_action(
    customers: int,
    paid: float,
    repeat: int,
    black: int,
    high: int = 0,
    paid_pct: float | None = None,
    customer_pct: float | None = None,
) -> str:
    if black > 0:
        return "本周给黑标预售素材和客户复盘，黑标客户由老板或核心运营接入一对一。"
    if paid_pct is not None and paid_pct >= 0.80:
        return "本周给专属成交复盘和下一批主推素材，验证高成交是否可复制到更多客户。"
    if repeat > 0:
        return "给复购承接素材，要求分销员逐个回访老客体验并记录二次需求。"
    if high > 0:
        return "先复盘高客单客户来源，给一版高信任解释素材，不急于扩大招募。"
    if customers > 0:
        return "给一个低门槛商品任务，先验证是否愿意持续转发和回收反馈。"
    return "只发基础招募/产品包，不投入一对一运营时间。"


def format_workbook(wb) -> None:
    header_fill = PatternFill("solid", fgColor="0B5C7E")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 8
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 42))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 36))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--detail", required=True, help="Core order-item detail export containing yz_open_id")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--week-label", required=True)
    parser.add_argument("--audit-note", help="Optional backend audit note markdown file")
    args = parser.parse_args()

    audit_note = Path(args.audit_note).expanduser() if args.audit_note else None
    output = build_tables(Path(args.detail).expanduser(), Path(args.output_dir).expanduser(), args.week_label, audit_note)
    print(output)


if __name__ == "__main__":
    main()
